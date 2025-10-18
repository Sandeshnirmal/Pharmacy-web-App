import openpyxl
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.contrib.auth import get_user_model
from product.models import Product, GenericName, Category, Composition, ProductComposition, Batch

User = get_user_model()

class Command(BaseCommand):
    help = 'Imports product data, including generic names, categories, compositions, and batches from an Excel file.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='The path to the Excel file (.xlsx) to import.')
        parser.add_argument('--created_by_id', type=int, help='ID of the user who created these products. Defaults to the first superuser found.')
        parser.add_argument('--sheet_name', type=str, default=None, help='Optional: Specify the name of the sheet to import from. Defaults to the active sheet.')

    def handle(self, *args, **options):
        excel_file_path = options['excel_file']
        created_by_id = options['created_by_id']
        sheet_name = options['sheet_name']

        try:
            if created_by_id:
                created_by_user = User.objects.get(id=created_by_id)
            else:
                created_by_user = User.objects.filter(is_superuser=True).first()
                if not created_by_user:
                    raise CommandError("No superuser found. Please specify --created_by_id or create a superuser.")
            self.stdout.write(self.style.SUCCESS(f"Products will be created by user: {created_by_user.username} (ID: {created_by_user.id})"))

        except User.DoesNotExist:
            raise CommandError(f"User with ID {created_by_id} does not exist.")
        except Exception as e:
            raise CommandError(f"Error determining created_by user: {e}")

        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise CommandError(f"Sheet '{sheet_name}' not found in the Excel file.")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            if not sheet.max_row > 1: # Check if there's at least a header and one data row
                raise CommandError("Excel sheet is empty or has no data rows.")

            headers = [cell.value.strip() if cell.value else '' for cell in sheet[1]]
            if not headers:
                raise CommandError("Excel sheet has no headers in the first row.")

            # Expected headers (case-insensitive for robustness)
            expected_headers = [
                'product_name', 'brand_name', 'generic_name', 'manufacturer',
                'medicine_type', 'prescription_type', 'strength', 'form',
                'dosage_form', 'pack_size', 'packaging_unit', 'description',
                'uses', 'side_effects', 'how_to_use', 'precautions', 'storage',
                'image_url', 'hsn_code', 'category_name', 'is_featured',
                'composition_name', 'composition_strength', 'composition_unit',
                'composition_percentage', 'composition_is_primary',
                'batch_number', 'manufacturing_date', 'expiry_date', 'quantity',
                'mrp_price', 'discount_percentage', 'min_stock_level'
            ]

            # Check if all expected headers are present
            missing_headers = [header for header in expected_headers if header not in headers]
            if missing_headers:
                self.stdout.write(self.style.WARNING(f"Missing some expected headers: {', '.join(missing_headers)}"))
                self.stdout.write(self.style.WARNING("Proceeding with available headers. Ensure your Excel file matches the expected format for full data import."))

            processed_products = 0
            for row_num in range(2, sheet.max_row + 1): # Start from 2 for data rows
                row_data = {headers[i]: sheet.cell(row=row_num, column=i+1).value for i in range(len(headers))}
                
                try:
                    with transaction.atomic():
                        # 1. Get or create GenericName
                        generic_name_str = row_data.get('generic_name')
                        if not generic_name_str:
                            self.stdout.write(self.style.ERROR(f"Skipping row {row_num}: 'generic_name' is required."))
                            continue
                        generic_name, _ = GenericName.objects.get_or_create(name=generic_name_str, defaults={'description': f"Generic name for {generic_name_str}"})

                        # 2. Get or create Category
                        category_name_str = row_data.get('category_name')
                        category = None
                        if category_name_str:
                            category, _ = Category.objects.get_or_create(name=category_name_str, defaults={'description': f"Category: {category_name_str}"})

                        # 3. Create or update Product
                        product_name = row_data.get('product_name')
                        manufacturer = row_data.get('manufacturer', 'MedCorp')
                        dosage_form = row_data.get('dosage_form', '')

                        if not product_name:
                            self.stdout.write(self.style.ERROR(f"Skipping row {row_num}: 'product_name' is required."))
                            continue

                        product_defaults = {
                            'brand_name': row_data.get('brand_name', ''),
                            'generic_name': generic_name,
                            'manufacturer': manufacturer,
                            'medicine_type': row_data.get('medicine_type', 'tablet'),
                            'prescription_type': row_data.get('prescription_type', 'otc'),
                            'strength': row_data.get('strength', ''),
                            'form': row_data.get('form', ''),
                            'is_prescription_required': row_data.get('prescription_type', 'otc').lower() != 'otc',
                            'min_stock_level': int(row_data.get('min_stock_level', 10) or 10),
                            'dosage_form': dosage_form,
                            'pack_size': row_data.get('pack_size', ''),
                            'packaging_unit': row_data.get('packaging_unit', ''),
                            'description': row_data.get('description', ''),
                            'uses': row_data.get('uses', ''),
                            'side_effects': row_data.get('side_effects', ''),
                            'how_to_use': row_data.get('how_to_use', ''),
                            'precautions': row_data.get('precautions', ''),
                            'storage': row_data.get('storage', ''),
                            'image_url': row_data.get('image_url', ''),
                            'hsn_code': row_data.get('hsn_code', ''),
                            'category': category,
                            'is_featured': str(row_data.get('is_featured', 'False')).lower() == 'true',
                            'created_by': created_by_user,
                        }

                        product, created = Product.objects.update_or_create(
                            name=product_name,
                            manufacturer=manufacturer,
                            dosage_form=dosage_form,
                            defaults=product_defaults
                        )

                        if created:
                            self.stdout.write(self.style.SUCCESS(f"Created Product: {product.name}"))
                        else:
                            self.stdout.write(self.style.WARNING(f"Updated Product: {product.name}"))

                        # 4. Handle Composition and ProductComposition (Many-to-Many)
                        composition_name_str = row_data.get('composition_name')
                        if composition_name_str:
                            composition, _ = Composition.objects.get_or_create(
                                name=composition_name_str,
                                defaults={'created_by': created_by_user}
                            )
                            product_composition_defaults = {
                                'strength': str(row_data.get('composition_strength', '')),
                                'unit': row_data.get('composition_unit', 'mg'),
                                'percentage': float(row_data['composition_percentage']) if row_data.get('composition_percentage') else None,
                                'is_primary': str(row_data.get('composition_is_primary', 'False')).lower() == 'true',
                            }
                            ProductComposition.objects.update_or_create(
                                product=product,
                                composition=composition,
                                defaults=product_composition_defaults
                            )
                            self.stdout.write(self.style.SUCCESS(f"  - Added/Updated Composition '{composition.name}' for {product.name}"))

                        # 5. Create Batch
                        batch_number = row_data.get('batch_number')
                        expiry_date_val = row_data.get('expiry_date')
                        quantity_val = row_data.get('quantity')
                        mrp_price_val = row_data.get('mrp_price')
                        discount_percentage_val = row_data.get('discount_percentage', 0)
                        manufacturing_date_val = row_data.get('manufacturing_date')

                        if batch_number and expiry_date_val and quantity_val is not None and mrp_price_val is not None:
                            try:
                                # openpyxl reads dates as datetime objects, convert to date
                                expiry_date = expiry_date_val.date() if isinstance(expiry_date_val, datetime) else datetime.strptime(str(expiry_date_val), '%Y-%m-%d').date()
                                manufacturing_date = manufacturing_date_val.date() if isinstance(manufacturing_date_val, datetime) else (datetime.strptime(str(manufacturing_date_val), '%Y-%m-%d').date() if manufacturing_date_val else None)
                                quantity = int(quantity_val)
                                mrp_price = float(mrp_price_val)
                                discount_percentage = float(discount_percentage_val)

                                batch, batch_created = Batch.objects.update_or_create(
                                    product=product,
                                    batch_number=batch_number,
                                    defaults={
                                        'manufacturing_date': manufacturing_date,
                                        'expiry_date': expiry_date,
                                        'quantity': quantity,
                                        'current_quantity': quantity, # Assuming current_quantity starts as total quantity
                                        'mrp_price': mrp_price,
                                        'discount_percentage': discount_percentage,
                                        'selling_price': mrp_price * (1 - discount_percentage / 100), # Calculated in save, but good to set here too
                                    }
                                )
                                if batch_created:
                                    self.stdout.write(self.style.SUCCESS(f"  - Created Batch '{batch.batch_number}' for {product.name}"))
                                else:
                                    self.stdout.write(self.style.WARNING(f"  - Updated Batch '{batch.batch_number}' for {product.name}"))
                            except ValueError as ve:
                                self.stdout.write(self.style.ERROR(f"  - Error processing batch data for {product.name} (row {row_num}): {ve}"))
                            except AttributeError as ae: # For date objects that might not be datetime
                                self.stdout.write(self.style.ERROR(f"  - Date format error for {product.name} (row {row_num}): {ae}. Ensure dates are in YYYY-MM-DD format or proper Excel date format."))
                        else:
                            self.stdout.write(self.style.WARNING(f"  - Skipping batch creation for {product.name} (row {row_num}): Missing batch_number, expiry_date, quantity, or mrp_price."))

                        processed_products += 1

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Error processing row {row_num}: {e} - Data: {row_data}"))
                    # Continue to next row even if one fails

            self.stdout.write(self.style.SUCCESS(f"Successfully processed {processed_products} product entries."))

        except FileNotFoundError:
            raise CommandError(f'File "{excel_file_path}" does not exist.')
        except Exception as e:
            raise CommandError(f"An unexpected error occurred: {e}")
