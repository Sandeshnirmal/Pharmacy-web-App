import os
import django
import openpyxl
from datetime import date
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from backend.product.models import Product, Composition, GenericName, Category, ProductComposition, Batch

User = get_user_model()

class Command(BaseCommand):
    help = 'Inserts product data from an Excel file into the database.'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the Excel file containing product data.')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except FileNotFoundError:
            raise CommandError(f"Error: Excel file '{file_path}' not found.")
        except Exception as e:
            raise CommandError(f"Error loading Excel file: {e}")

        headers = [cell.value for cell in sheet[1]]
        
        try:
            default_user = User.objects.get(id=1)
        except User.DoesNotExist:
            raise CommandError("Error: Default user with ID 1 not found. Please create a user or specify an existing user ID.")

        for row_index in range(2, sheet.max_row + 1):
            row_data = {headers[i]: sheet.cell(row=row_index, column=i+1).value for i in range(len(headers))}

            self.stdout.write(f"Processing product: {row_data.get('product_name')}")

            # Get or create GenericName
            generic_name_obj, _ = GenericName.objects.get_or_create(
                name=row_data['generic_name'],
                defaults={'description': f"Generic name for {row_data['generic_name']}"}
            )

            # Get or create Category
            category_obj, _ = Category.objects.get_or_create(
                name=row_data['category_name'],
                defaults={'description': f"Category for {row_data['category_name']}"}
            )

            # Get or create Composition
            composition_obj, _ = Composition.objects.get_or_create(
                name=row_data['composition_name'],
                defaults={
                    'description': f"Composition: {row_data['composition_name']}",
                    'created_by': default_user
                }
            )

            # Create or update Product
            product_defaults = {
                'brand_name': row_data['brand_name'],
                'generic_name': generic_name_obj,
                'manufacturer': row_data['manufacturer'],
                'medicine_type': row_data['medicine_type'],
                'prescription_type': row_data['prescription_type'],
                'strength': row_data['strength'],
                'form': row_data['form'],
                'min_stock_level': int(row_data['min_stock_level']),
                'dosage_form': row_data['dosage_form'],
                'pack_size': row_data['pack_size'],
                'packaging_unit': row_data['packaging_unit'],
                'description': row_data['description'],
                'uses': row_data['uses'],
                'side_effects': row_data['side_effects'],
                'how_to_use': row_data['how_to_use'],
                'precautions': row_data['precautions'],
                'storage': row_data['storage'],
                'image_url': row_data['image_url'],
                'hsn_code': row_data['hsn_code'],
                'category': category_obj,
                'is_featured': row_data['is_featured'].lower() == 'true',
                'is_prescription_required': row_data['prescription_type'].lower() == 'rx',
                'created_by': default_user,
            }
            
            product, created = Product.objects.update_or_create(
                name=row_data['product_name'],
                manufacturer=row_data['manufacturer'],
                dosage_form=row_data['dosage_form'],
                defaults=product_defaults
            )

            if created:
                self.stdout.write(self.style.SUCCESS(f"Created new product: {product.name}"))
            else:
                self.stdout.write(f"Updated existing product: {product.name}")

            # Create ProductComposition
            ProductComposition.objects.update_or_create(
                product=product,
                composition=composition_obj,
                defaults={
                    'strength': row_data['composition_strength'],
                    'unit': row_data['composition_unit'],
                    'percentage': Decimal(row_data['composition_percentage']),
                    'is_primary': row_data['composition_is_primary'].lower() == 'true',
                }
            )

            # Create Batch
            Batch.objects.update_or_create(
                product=product,
                batch_number=row_data['batch_number'],
                defaults={
                    'manufacturing_date': row_data['manufacturing_date'],
                    'expiry_date': row_data['expiry_date'],
                    'quantity': int(row_data['quantity']),
                    'current_quantity': int(row_data['quantity']),
                    'mrp_price': Decimal(row_data['mrp_price']),
                    'discount_percentage': Decimal(row_data['discount_percentage']),
                    'selling_price': Decimal(row_data['mrp_price']) * (1 - Decimal(row_data['discount_percentage']) / 100),
                    'online_mrp_price': Decimal(row_data['mrp_price']),
                    'online_discount_percentage': Decimal(row_data['discount_percentage']),
                    'online_selling_price': Decimal(row_data['mrp_price']) * (1 - Decimal(row_data['discount_percentage']) / 100),
                    'offline_mrp_price': Decimal(row_data['mrp_price']),
                    'offline_discount_percentage': Decimal(row_data['discount_percentage']),
                    'offline_selling_price': Decimal(row_data['mrp_price']) * (1 - Decimal(row_data['discount_percentage']) / 100),
                }
            )
            self.stdout.write(f"Successfully processed product: {product.name} with batch {row_data['batch_number']}")

        self.stdout.write(self.style.SUCCESS("\nData insertion complete."))
